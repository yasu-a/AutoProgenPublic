from collections import Counter
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from functools import cached_property, cache
from pathlib import Path
from typing import Any, Generator

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from domain.models.student_mark import StudentMark
from domain.models.values import StudentID, TargetID
from utils.app_logging import create_logger


class ScoreExcelError(RuntimeError):
    pass


class ScoreExcelMalformedWorksheetError(ScoreExcelError):
    def __init__(self, *, message: str):
        self._message = message

    @property
    def message(self):
        return self._message


@dataclass
class _ScoreWritingPosition:
    student_id: StudentID
    target_id: TargetID
    row_index: int
    column_index: int


class _ScoreExcelWorksheetReader:
    def __init__(self, sheet: Worksheet):
        self._ws = sheet
        self._rows_str = tuple(
            tuple(
                None if v is None else str(v)
                for v in row
            ) for row in self._ws.values
        )

    @cached_property
    def _table_header(self) -> int:
        for i_row, row in enumerate(self._rows_str):
            if row[0] is not None and row[0].strip().startswith("#"):
                return i_row
        raise ScoreExcelMalformedWorksheetError(
            message="ワークシートに「# 学籍番号」「# 氏名」「# 合計点」「問1」・・・で始まるヘッダーが見つかりません",
        )

    @cached_property
    def _table_begin(self) -> int:
        return self._table_header + 1

    @cache
    def _try_get_student_id(self, i_row: int) -> StudentID | None:
        row = self._rows_str[i_row]
        if row[0] is None:
            return None
        try:
            student_id = StudentID(row[0].strip())
        except ValueError:
            return None
        else:
            return student_id

    @cached_property
    def _table_end(self) -> int:  # exclusive
        i_row = self._table_begin
        while i_row < len(self._rows_str):
            if self._try_get_student_id(i_row) is None:
                break
            i_row += 1
        return i_row

    @cache
    def _find_row_index_of_student_id(self, student_id: StudentID) -> int:
        for i_row in range(self._table_begin, self._table_end):
            student_id_of_row = self._try_get_student_id(i_row)
            if student_id == student_id_of_row:
                return i_row
        raise ScoreExcelMalformedWorksheetError(
            message=f"ワークシートに学籍番号{student_id}の行が見つかりません",
        )

    @cache
    def _find_column_index_of_target_id(self, target_id: TargetID) -> int:
        target_number = int(target_id)
        header = self._rows_str[self._table_header]
        for i_column, title in enumerate(header):
            if title.strip() == f"問{target_number}":
                return i_column
        raise ScoreExcelMalformedWorksheetError(
            message=f"ワークシートに問{target_number}の列が見つかりません",
        )

    def _validate_header(self):
        header = self._rows_str[self._table_header]
        if tuple(header[:4]) != ("# 学籍番号", "# 氏名", "# 合計点", "問1"):
            raise ScoreExcelMalformedWorksheetError(
                message=(
                        f"ヘッダーは「# 学籍番号」「# 氏名」「# 合計点」「問1」・・・で始まる必要があります: "
                        + " ".join(map(str, header[:4]))
                ),
            )

    def _validate_student_ids(self, student_ids: list[StudentID]):
        # アプリケーションから渡された学籍番号の集合を生成
        source_student_id_set = set(student_ids)

        # ワークシート上の学籍番号の集合を生成
        worksheet_student_id_lst = [
            self._try_get_student_id(i)
            for i in range(self._table_begin, self._table_end)
        ]
        worksheet_student_id_set = set(worksheet_student_id_lst)
        if len(worksheet_student_id_set) != len(worksheet_student_id_lst):
            duplicated_student_ids = [
                student_id
                for student_id, count in Counter(worksheet_student_id_lst)
                if count >= 2
            ]
            raise ScoreExcelMalformedWorksheetError(
                message=(
                        "ワークシートの学籍番号が重複しています: "
                        + " ".join(map(str, duplicated_student_ids))
                ),
            )

        # 差を検証
        source_only = source_student_id_set - worksheet_student_id_set
        worksheet_only = worksheet_student_id_set - source_student_id_set
        if source_only or worksheet_only:
            messages = ["学籍番号が一致しません"]
            if source_only:
                messages.append(
                    "ワークシートに存在しません: "
                    + " ".join(map(str, source_only))
                )
            if worksheet_only:
                messages.append(
                    "ワークシートに存在しますがプロジェクトに存在しません: "
                    + " ".join(map(str, worksheet_only))
                )
            raise ScoreExcelMalformedWorksheetError(
                message=" ".join(messages)
            )

    def validate(self, student_ids: list[StudentID]):
        self._validate_header()
        self._validate_student_ids(student_ids)

    def get_writing_positions(self, student_ids: list[StudentID], target_id: TargetID) \
            -> dict[StudentID, _ScoreWritingPosition]:  # i_row, i_column
        pos_lst: dict[StudentID, _ScoreWritingPosition] = {}
        for student_id in student_ids:
            pos = _ScoreWritingPosition(
                student_id=student_id,
                target_id=target_id,
                row_index=self._find_row_index_of_student_id(student_id),
                column_index=self._find_column_index_of_target_id(target_id),
            )
            pos_lst[student_id] = pos
        return pos_lst


@dataclass(frozen=True)
class WorksheetStat:
    name: str
    valid: bool
    message: str | None


class ScoreExcelIO:
    _logger = create_logger()

    def __init__(self, *, excel_fullpath: Path):
        self._excel_fullpath = excel_fullpath

    @contextmanager
    def _open(self, *, readonly: bool) -> Generator[Workbook, Any, None]:
        wb = openpyxl.open(self._excel_fullpath, read_only=readonly)
        try:
            yield wb
        finally:
            wb.close()

    def list_worksheet_stats(
            self,
            *,
            student_ids: list[StudentID],
    ) -> list[WorksheetStat]:
        with self._open(readonly=True) as wb:
            ws_stats = []
            for ws_name in wb.sheetnames:
                ws: Worksheet = wb[ws_name]
                try:
                    _ScoreExcelWorksheetReader(ws).validate(student_ids)
                except ScoreExcelMalformedWorksheetError as e:
                    self._logger.exception(f"Invalid worksheet: {ws_name}")
                    ws_stat = WorksheetStat(
                        name=ws_name,
                        valid=False,
                        message=e.message,
                    )
                else:
                    self._logger.debug(f"Valid worksheet: {ws_name}")
                    ws_stat = WorksheetStat(
                        name=ws_name,
                        valid=True,
                        message=None,
                    )
                ws_stats.append(ws_stat)
            return ws_stats

    def has_data(
            self,
            *,
            worksheet_name: str,
            student_ids: list[StudentID],
            target_id: TargetID,
    ):
        with self._open(readonly=True) as wb:
            ws = wb[worksheet_name]
            reader = _ScoreExcelWorksheetReader(ws)
            writing_positions = reader.get_writing_positions(student_ids, target_id)
            for student_id in student_ids:
                pos = writing_positions[student_id]
                cell = ws.cell(row=pos.row_index + 1, column=pos.column_index + 1)
                if cell.value is not None:
                    return True
            return False

    def apply(
            self,
            *,
            worksheet_name: str,
            student_marks: list[StudentMark],
            target_id: TargetID,
            do_backup: bool,
    ) -> Path:  # backup path
        student_ids = [student_mark.student_id for student_mark in student_marks]
        with self._open(readonly=False) as wb:
            if do_backup:
                timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
                backup_path = self._excel_fullpath.parent / (
                        self._excel_fullpath.stem + f"-backup-{timestamp}.xlsx")
                wb.save(backup_path)
            else:
                backup_path = None

            ws = wb[worksheet_name]
            reader = _ScoreExcelWorksheetReader(ws)
            writing_positions = reader.get_writing_positions(student_ids, target_id)
            for student_mark in student_marks:
                pos = writing_positions[student_mark.student_id]
                if student_mark.is_marked:
                    ws.cell(
                        row=pos.row_index + 1,
                        column=pos.column_index + 1,
                        value=student_mark.score,
                    )
                else:
                    ws.cell(
                        row=pos.row_index + 1,
                        column=pos.column_index + 1,
                        value="",
                    )

            try:
                wb.save(self._excel_fullpath)
            except PermissionError:
                raise ScoreExcelError("書き込みを拒否されました。ファイルをExcelで開いていませんか？")

        return backup_path
