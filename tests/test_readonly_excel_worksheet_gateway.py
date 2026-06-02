import io

import openpyxl
import pytest

from domain.error import ReadonlyExcelWorksheetGatewayError
from infra.gateway.readonly_excel_worksheet import ReadonlyExcelWorksheetGateway


def _workbook_to_bytes(wb: openpyxl.Workbook) -> bytes:
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def test_readonly_excel_worksheet_gateway_normalizes_empty_cell_and_hyperlink() -> None:
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws["A1"] = "hello"
    ws["B1"] = None
    ws["C1"] = "開く"
    ws["C1"].hyperlink = "21D5109047B@21D5109047B\\"

    excel_bytes = _workbook_to_bytes(wb)
    worksheet = ReadonlyExcelWorksheetGateway().read_from_bytes(excel_bytes=excel_bytes)

    assert worksheet.cell_at(row_index=0, column_index=0).text == "hello"
    assert worksheet.cell_at(row_index=0, column_index=1).text == ""
    assert worksheet.cell_at(row_index=0, column_index=1).hyperlink_target is None
    assert worksheet.cell_at(row_index=0, column_index=2).hyperlink_target == "21D5109047B@21D5109047B\\"


def test_readonly_excel_worksheet_gateway_rejects_multiple_worksheets() -> None:
    wb = openpyxl.Workbook()
    wb.create_sheet("second")
    excel_bytes = _workbook_to_bytes(wb)

    with pytest.raises(ReadonlyExcelWorksheetGatewayError) as exc_info:
        ReadonlyExcelWorksheetGateway().read_from_bytes(excel_bytes=excel_bytes)

    assert "ワークブックは1つだけワークシート" in exc_info.value.reason
