from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from shared.domain.interface.gateway import (
    IExcelGateway,
    ExcelGatewayError,
)
from shared.domain.value.excel_cell_table import ExcelCellTable


class ExcelGateway(IExcelGateway):
    """Excel読み書きGatewayの実装"""
    
    def list_sheet_names(self, excel_path: Path) -> list[str]:
        """シート名の一覧を取得"""
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            names = wb.sheetnames
            wb.close()
            return names
        except Exception as e:
            raise ExcelGatewayError(f"Excelファイルの読み込みに失敗しました: {e}") from e
    
    def get_sheet_cells(
        self,
        excel_path: Path,
        sheet_name: str,
        *,
        max_rows: int | None = None,
        max_cols: int | None = None,
    ) -> ExcelCellTable:
        """
        シートのセルを取得
        
        Args:
            excel_path: Excelファイルのパス
            sheet_name: シート名
            max_rows: 読み込む最大行数（Noneの場合は全行）
            max_cols: 読み込む最大列数（Noneの場合は全列）
        
        Returns:
            ExcelCellTable: ExcelセルテーブルのValue Object
        """
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            try:
                if sheet_name not in wb.sheetnames:
                    raise ExcelGatewayError(f"シート '{sheet_name}' が見つかりません")
                
                ws: Worksheet = wb[sheet_name]
                data = {}
                # 読み込み範囲を決定
                actual_max_row = min(ws.max_row, max_rows) if max_rows is not None else ws.max_row
                actual_max_col = min(ws.max_column, max_cols) if max_cols is not None else ws.max_column
                
                # iter_rowsで一気に読み込む（values_only=Trueで値のみ取得）
                for row_idx, row in enumerate(ws.iter_rows(
                    min_row=1,
                    max_row=actual_max_row,
                    min_col=1,
                    max_col=actual_max_col,
                    values_only=True
                )):
                    for col_idx, cell_value in enumerate(row):
                        # セルの値がNoneまたは空の場合は空文字列として返す
                        if cell_value is None:
                            data[(row_idx, col_idx)] = ""
                        else:
                            val_str = str(cell_value)
                            # 数値の0は空文字列として表示（Excelの見た目に合わせる）
                            if val_str == "0" and isinstance(cell_value, (int, float)):
                                data[(row_idx, col_idx)] = ""
                            else:
                                data[(row_idx, col_idx)] = val_str
                return ExcelCellTable(data=data)
            finally:
                wb.close()
        except ExcelGatewayError:
            raise
        except Exception as e:
            raise ExcelGatewayError(f"シートの読み込みに失敗しました: {e}") from e
    
    def update_sheet_cells(
        self,
        *,
        excel_path: Path,
        sheet_name: str,
        values: dict[tuple[int, int], Any],
    ) -> None:
        """
        シートのセルを更新
        
        Args:
            excel_path: Excelファイルのパス
            sheet_name: シート名
            values: (row_index, column_index) -> value のマッピング
                row_index, column_indexは0-based
        """
        try:
            wb: Workbook = openpyxl.load_workbook(excel_path, read_only=False)
            try:
                if sheet_name not in wb.sheetnames:
                    raise ExcelGatewayError(f"シート '{sheet_name}' が見つかりません")
                
                ws: Worksheet = wb[sheet_name]
                
                # セルを更新（openpyxlは1-basedなので+1する）
                for (row_index, col_index), value in values.items():
                    ws.cell(row=row_index + 1, column=col_index + 1, value=value)
                
                wb.save(excel_path)
            finally:
                wb.close()
        except ExcelGatewayError:
            raise
        except PermissionError:
            raise ExcelGatewayError("書き込みを拒否されました。ファイルをExcelで開いていませんか？")
        except Exception as e:
            raise ExcelGatewayError(f"シートの更新に失敗しました: {e}") from e

