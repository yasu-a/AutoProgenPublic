from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from shared.domain.interface.gateway import (
    IExcelScoreExportGateway,
    ExcelScoreExportGatewayError,
)


class ExcelScoreExportGateway(IExcelScoreExportGateway):
    """Excel読み書きGatewayの実装"""
    
    def list_sheet_names(self, excel_path: Path) -> list[str]:
        """シート名の一覧を取得"""
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            names = wb.sheetnames
            wb.close()
            return names
        except Exception as e:
            raise ExcelScoreExportGatewayError(f"Excelファイルの読み込みに失敗しました: {e}") from e
    
    def get_sheet_cells(self, excel_path: Path, sheet_name: str) -> dict[tuple[int, int], str]:
        """
        シートの全セルを取得
        
        Returns:
            dict[tuple[int, int], str]: (row_index, column_index) -> cell_value のマッピング
            row_index, column_indexは0-based
            Noneのセルは空文字""として返す
        """
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            try:
                if sheet_name not in wb.sheetnames:
                    raise ExcelScoreExportGatewayError(f"シート '{sheet_name}' が見つかりません")
                
                ws: Worksheet = wb[sheet_name]
                data = {}
                # プレビュー用に全セル読み込み (0-based indexで格納)
                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    for col_idx, val in enumerate(row):
                        if val is not None:
                            data[(row_idx, col_idx)] = str(val)
                return data
            finally:
                wb.close()
        except ExcelScoreExportGatewayError:
            raise
        except Exception as e:
            raise ExcelScoreExportGatewayError(f"シートの読み込みに失敗しました: {e}") from e
    
    def update_sheet_cells(
        self,
        *,
        excel_path: Path,
        sheet_name: str,
        values: dict[tuple[int, int], Any],
    ) -> None:
        """
        シートのセルを更新
        
        Args:
            excel_path: Excelファイルのパス
            sheet_name: シート名
            values: (row_index, column_index) -> value のマッピング
                row_index, column_indexは0-based
        """
        try:
            wb: Workbook = openpyxl.load_workbook(excel_path, read_only=False)
            try:
                if sheet_name not in wb.sheetnames:
                    raise ExcelScoreExportGatewayError(f"シート '{sheet_name}' が見つかりません")
                
                ws: Worksheet = wb[sheet_name]
                
                # セルを更新（openpyxlは1-basedなので+1する）
                for (row_index, col_index), value in values.items():
                    ws.cell(row=row_index + 1, column=col_index + 1, value=value)
                
                wb.save(excel_path)
            finally:
                wb.close()
        except ExcelScoreExportGatewayError:
            raise
        except PermissionError:
            raise ExcelScoreExportGatewayError("書き込みを拒否されました。ファイルをExcelで開いていませんか？")
        except Exception as e:
            raise ExcelScoreExportGatewayError(f"シートの更新に失敗しました: {e}") from e

